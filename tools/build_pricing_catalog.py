#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import shutil
from io import BytesIO
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from PIL import Image


ROOT = Path("/Users/lehjke/Desktop/Dokumentatsia")
OUT = Path("src/TFlexDrawingService.Api/Data/pricing-catalog.json")
SMEC_ASSET_OUT = Path("src/TFlexDrawingService.Api/wwwroot/assets/smec")
SMEC_FORM_EXAMPLE = Path("/Users/lehjke/Downloads/Example_sp.xlsx")
if not SMEC_FORM_EXAMPLE.exists():
    SMEC_FORM_EXAMPLE = ROOT / "P240174-METEOR-Версаль.xlsx"
XIZI_FORM_EXAMPLE = ROOT / "Spetsifikatsia_XIZI_v_0_2_1.xlsx"
XIZI_FINISHES = ROOT / "Otdelki_XIZI.xlsx"
XIZI_ASSET_OUT = Path("src/TFlexDrawingService.Api/wwwroot/assets/xizi")


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def is_excel_stop_marker(value: Any) -> bool:
    return isinstance(value, str) and set(value.strip()) == {"\\"}


def as_number(value: Any) -> float | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("¥", "").replace(",", ".").strip()
    try:
        return float(text)
    except ValueError:
        return None


def parse_xizi_base(wb) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for sheet in ["UN-Victor R", "UN-Victor MRL", "G3"]:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        starts = [
            index
            for index, value in enumerate(rows[1])
            if clean(value) == "Stops / Capacity"
        ]
        for start in starts:
            extra_rise_row = next(
                (
                    row
                    for row in rows
                    if start < len(row) and clean(row[start]) == "Extra rise"
                ),
                (),
            )
            door_opening_row = next(
                (
                    row
                    for row in rows
                    if start < len(row) and clean(row[start]) == "Door opening"
                ),
                (),
            )
            for column in range(start + 1, min(start + 10, len(rows[1]))):
                capacity = as_number(rows[1][column])
                speed = as_number(rows[0][column])
                if capacity is None or speed is None:
                    continue
                extra_rise = clean(extra_rise_row[column] if column < len(extra_rise_row) else None)
                door_opening = clean(door_opening_row[column] if column < len(door_opening_row) else None)
                for row in rows[2:]:
                    stops = as_number(row[start] if start < len(row) else None)
                    price = clean(row[column] if column < len(row) else None)
                    if stops is None:
                        continue
                    entries.append(
                        {
                            "series": sheet,
                            "capacity": int(capacity),
                            "speed": speed,
                            "stops": int(stops),
                            "price": price,
                            "extraRisePerMeter": extra_rise,
                            "standardDoorOpening": door_opening,
                        }
                    )
    return entries


def parse_xizi_doors(wb) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for sheet in ["Doors FERMATOR", "Doors OPTIMAX"]:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        manufacturer = sheet.replace("Doors ", "")
        current_part = None
        current_type = None
        current_fire = None
        current_finish = None
        row_index = 0
        while row_index < len(rows):
            row = rows[row_index]
            first = clean(row[0] if len(row) > 0 else None)
            second = clean(row[1] if len(row) > 1 else None)
            if first in {"Shaft door", "Car door", "2nd door"}:
                current_part = first
                current_type = None
                current_fire = None
                current_finish = None
                row_index += 1
                continue
            if first in {"2S", "CO"}:
                current_type = first
                current_fire = None
                current_finish = None
                row_index += 1
                continue
            if first in {"None", "E30", "EI60"}:
                current_fire = first
                row_index += 1
                continue
            if first in {"Painted steel", "AISI443", "SUS HL", "SUS Mirror", "Glass"}:
                current_finish = first
                row_index += 1
                continue
            if first == "CAP" and second == "Floor":
                width_row = rows[row_index + 1] if row_index + 1 < len(rows) else ()
                widths = [as_number(value) for value in width_row[2:]]
                data_row = row_index + 2
                while data_row < len(rows):
                    data = rows[data_row]
                    cap = as_number(data[0] if len(data) > 0 else None)
                    floor = clean(data[1] if len(data) > 1 else None)
                    if cap is None or floor is None:
                        break
                    for offset, width in enumerate(widths, start=2):
                        if width is None:
                            continue
                        entries.append(
                            {
                                "manufacturer": manufacturer,
                                "part": current_part,
                                "doorType": current_type,
                                "fireRating": current_fire,
                                "finish": current_finish,
                                "capacity": int(cap),
                                "floor": str(floor),
                                "width": int(width),
                                "price": clean(data[offset] if offset < len(data) else None),
                            }
                        )
                    data_row += 1
                row_index = data_row
                continue
            row_index += 1
    return entries


def parse_xizi_simple_price_sheet(wb, sheet: str, category: str) -> list[dict[str, Any]]:
    ws = wb[sheet]
    entries: list[dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        code = clean(row[0] if len(row) > 0 else None)
        prices = [clean(value) for value in row[1:] if clean(value) is not None]
        price = prices[0] if prices else None
        if not code or code in {category, "#"}:
            continue
        if as_number(price) is None and price != -1:
            continue
        entries.append(
            {
                "category": category,
                "code": str(code).strip(),
                "price": price,
                "prices": prices,
            }
        )
    return entries


def parse_xizi_option_descriptions(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, data_only=True, read_only=False)
    worksheet = workbook["L1"] if "L1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    descriptions: dict[str, str] = {}
    for row in range(28, min(worksheet.max_row, 51) + 1):
        code = clean(worksheet.cell(row, 2).value)
        description = clean(worksheet.cell(row, 4).value)
        if code and description:
            descriptions[str(code).strip()] = str(description).strip()
    return descriptions


def enrich_xizi_options(entries: list[dict[str, Any]], descriptions: dict[str, str]) -> list[dict[str, Any]]:
    for entry in entries:
        description = descriptions.get(entry["code"])
        if description:
            entry["description"] = description
    return entries


def parse_xizi_decoration(wb) -> list[dict[str, Any]]:
    ws = wb["Decoration"]
    rows = list(ws.iter_rows(values_only=True))
    entries: list[dict[str, Any]] = []
    row = 0
    while row < len(rows):
        section = clean(rows[row][0] if rows[row] else None)
        next_first = clean(rows[row + 1][0] if row + 1 < len(rows) and rows[row + 1] else None)
        if section and next_first == "#" and row + 3 < len(rows):
            codes = [clean(value) for value in rows[row + 1][1:]]
            value_label = clean(rows[row + 2][0] if rows[row + 2] else None)
            has_height_row = value_label == "CH"
            heights = [clean(value) for value in rows[row + 2][1:]] if has_height_row else [None] * len(codes)
            price_row = row + 3 if has_height_row else row + 2
            overprice_row = price_row + 1
            prices = [clean(value) for value in rows[price_row][1:]]
            overprices = [
                clean(value)
                for value in (rows[overprice_row][1:] if overprice_row < len(rows) else ())
            ]
            for index, code in enumerate(codes):
                if not code:
                    continue
                entries.append(
                    {
                        "category": str(section),
                        "code": str(code),
                        "height": heights[index] if index < len(heights) else None,
                        "price": prices[index] if index < len(prices) else None,
                        "overprice": overprices[index] if index < len(overprices) else None,
                    }
                )
            row = overprice_row + 1
        else:
            row += 1
    return entries


def parse_xizi_finish_descriptions(wb) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_column > 30 or ws.max_row > 500:
            continue
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                value = clean(ws.cell(row, col).value)
                if isinstance(value, str) and value.startswith(("U-", "Z")):
                    description = clean(ws.cell(row + 1, col).value) or clean(ws.cell(row, col + 1).value)
                    entries.append(
                        {
                            "category": sheet.strip(),
                            "code": value,
                            "description": description,
                        }
                    )
    return entries


def normalize_xizi_code(value: Any) -> str | None:
    code = clean(value)
    if code is None:
        return None
    return str(code).translate(str.maketrans({"С": "C", "А": "A", "В": "B", "Е": "E"})).strip()


def xizi_image_metadata(ws, row: int, column: int) -> tuple[str | None, str | None]:
    sheet = ws.title.strip()
    code: Any = None
    descriptions: list[Any] = []

    if sheet == "Кабины":
        code_row = 1 if row == 2 else 12 if row == 13 else 23 if row == 24 else row - 1
        code = ws.cell(code_row, column).value
        descriptions = [ws.cell(index, column).value for index in range(code_row + 2, min(code_row + 9, ws.max_row) + 1)]
    elif sheet == "Пол":
        code_column = 2 if column == 3 else 5
        code = ws.cell(row, code_column).value
        descriptions = [ws.cell(2, column).value]
    elif sheet == "Потолок":
        code_column = 2 if column == 4 else 6
        description_column = 3 if column == 4 else 7
        code = ws.cell(row, code_column).value
        descriptions = [ws.cell(row, description_column).value]
    elif sheet == "Поручни":
        code = ws.cell(1, column).value
        descriptions = [ws.cell(3, column).value]
    elif sheet == "Вызывные посты":
        code_row = 1 if row == 2 else 5 if row == 6 else 9
        code = ws.cell(code_row, column).value
        descriptions = [ws.cell(code_row + 2, column).value]
    elif sheet == "Индикаторы положения кабины":
        code = ws.cell(1, column).value
        descriptions = [ws.cell(3, column).value]
    elif sheet == "COP":
        code = ws.cell(1, column).value
        descriptions = [ws.cell(3, column).value]
        if column == 9 and normalize_xizi_code(code) == "U-CF100C":
            code = "U-CF2400C"
    elif sheet == "DDS":
        code = ws.cell(1, column).value

    normalized_code = normalize_xizi_code(code)
    description = " · ".join(
        dict.fromkeys(str(value).strip() for value in descriptions if clean(value) is not None)
    )
    return normalized_code, description or None


def xizi_code_aliases(code: str) -> list[str]:
    aliases = [code]
    if code.startswith("U-ZW3400-"):
        aliases.append(code.replace("U-ZW3400-", "U-ZW3410-", 1))
    if code == "U-ZW1600-f":
        aliases.append("U-ZW1600-F")
    if code == "U-HW100":
        aliases.extend(['U-HW100(7_TFT)', 'U-HW100(7_BND)', 'U-HW100(7_LED)'])
    if code == "U-HW200":
        aliases.extend(['U-HW200(7_TFT)', 'U-HW200(7_LED)'])
    return aliases


def write_xizi_image(image_data: bytes, code: str) -> str:
    safe_name = re.sub(r"[^a-zA-Z0-9._-]+", "-", code).strip("-").lower()
    output = XIZI_ASSET_OUT / f"{safe_name}.webp"
    with Image.open(BytesIO(image_data)) as source:
        source.thumbnail((640, 640), Image.Resampling.LANCZOS)
        converted = source.convert("RGB")
        converted.save(output, "WEBP", quality=84, method=6)
    return f"/assets/xizi/{output.name}"


def copy_xizi_finish_images() -> tuple[list[dict[str, Any]], dict[str, str], dict[str, str]]:
    if not XIZI_FINISHES.exists():
        return [], {}, {}

    if XIZI_ASSET_OUT.exists():
        shutil.rmtree(XIZI_ASSET_OUT)
    XIZI_ASSET_OUT.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(XIZI_FINISHES, data_only=True, read_only=False)
    visual_items: dict[str, dict[str, Any]] = {}
    image_map: dict[str, str] = {}
    description_map: dict[str, str] = {}

    for worksheet in workbook.worksheets:
        images = sorted(
            worksheet._images,
            key=lambda image: (image.anchor._from.row, image.anchor._from.col),
        )
        for image in images:
            row = image.anchor._from.row + 1
            column = image.anchor._from.col + 1
            code, description = xizi_image_metadata(worksheet, row, column)
            if not code:
                continue

            aliases = xizi_code_aliases(code)
            existing = next((image_map[alias] for alias in aliases if alias in image_map), None)
            image_url = existing or write_xizi_image(image._data(), code)
            for alias in aliases:
                if alias in visual_items:
                    continue
                visual_items[alias] = {
                    "code": alias,
                    "imageUrl": image_url,
                    "description": description,
                }
                image_map[alias] = image_url
                if description:
                    description_map[alias] = description

    return list(visual_items.values()), image_map, description_map


def enrich_xizi_visual_entries(
    entries: list[dict[str, Any]],
    image_map: dict[str, str],
    description_map: dict[str, str],
) -> list[dict[str, Any]]:
    for entry in entries:
        code = normalize_xizi_code(entry.get("code"))
        if not code:
            continue
        entry["code"] = code
        if code in image_map:
            entry["imageUrl"] = image_map[code]
        if code in description_map and not entry.get("description"):
            entry["description"] = description_map[code]
    return entries


def parse_xizi_containers(wb) -> list[dict[str, Any]]:
    ws = wb["40HQ"]
    rows = list(ws.iter_rows(values_only=True))
    entries: list[dict[str, Any]] = []
    capacities = [as_number(value) for value in rows[0][1:]]
    for row in rows[1:]:
        stops = as_number(row[0] if row else None)
        if stops is None:
            continue
        for offset, capacity in enumerate(capacities, start=1):
            if capacity is None:
                continue
            fraction = clean(row[offset] if offset < len(row) else None)
            entries.append(
                {
                    "capacity": int(capacity),
                    "stops": int(stops),
                    "container": "40HQ",
                    "fraction": fraction,
                }
            )
    return entries


def read_validation_choices(path: Path, sheet_name: str, field_names: dict[str, str]) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    workbook_formulas = load_workbook(path, data_only=False, read_only=False)
    workbook_values = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in workbook_formulas.sheetnames:
        sheet_name = "P1" if "P1" in workbook_formulas.sheetnames else workbook_formulas.sheetnames[0]
    ws_formulas = workbook_formulas[sheet_name]
    ws_values = workbook_values[sheet_name]

    def cell_value(token: str) -> Any:
        row, column = coordinate_to_tuple(token.replace("$", ""))
        return clean(ws_values.cell(row, column).value)

    def values_from_range(reference: str) -> list[Any]:
        min_col, min_row, max_col, max_row = range_boundaries(reference.replace("$", ""))
        values: list[Any] = []
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                value = clean(ws_values.cell(row, column).value)
                if value is not None and value != "\\":
                    values.append(value)
        return values

    def eval_argument(argument: str) -> int:
        argument = argument.strip()
        if not argument:
            return 0
        if re.fullmatch(r"\$?[A-Z]+\$?\d+", argument):
            value = cell_value(argument)
        else:
            value = argument
        try:
            return int(float(str(value).replace(",", ".")))
        except (TypeError, ValueError):
            return 0

    def values_from_formula(formula: str | None) -> list[Any]:
        if not formula:
            return []
        formula = formula.strip()
        if formula.startswith('"') and formula.endswith('"'):
            return [item.strip() for item in formula[1:-1].split(",") if item.strip()]
        if ":" in formula and not formula.upper().startswith("OFFSET("):
            return values_from_range(formula)

        match = re.match(r"OFFSET\(([^,]+),([^,]*),([^,]*),([^,]*),([^)]*)\)", formula, re.IGNORECASE)
        if not match:
            return []

        base, row_offset, column_offset, height, width = [part.strip() for part in match.groups()]
        base_row, base_column = coordinate_to_tuple(base.replace("$", ""))
        min_row = base_row + eval_argument(row_offset)
        min_column = base_column + eval_argument(column_offset)
        row_count = max(1, eval_argument(height))
        column_count = max(1, eval_argument(width))
        values: list[Any] = []
        for row in range(min_row, min_row + row_count):
            for column in range(min_column, min_column + column_count):
                value = clean(ws_values.cell(row, column).value)
                if value is not None and value != "\\":
                    values.append(value)
        return values

    groups_by_name: dict[str, dict[str, Any]] = {}
    for validation in ws_formulas.data_validations.dataValidation:
        if validation.type != "list":
            continue
        options = values_from_formula(validation.formula1)
        if not options:
            continue
        first_cell = str(validation.sqref).split()[0].split(":")[0]
        name = field_names.get(first_cell, first_cell)
        existing = groups_by_name.setdefault(
            name,
            {
                "name": name,
                "sourceSheet": sheet_name,
                "cells": [],
                "options": [],
            },
        )
        existing["cells"].extend(str(validation.sqref).split())
        for option in options:
            normalized_option = str(option)
            if normalized_option not in existing["options"]:
                existing["options"].append(normalized_option)

    return sorted(groups_by_name.values(), key=lambda item: item["name"])


def parse_xizi_choice_groups(path: Path) -> list[dict[str, Any]]:
    return read_validation_choices(
        path,
        "L1",
        {
            "C6": "Elevator Type",
            "F6": "Model",
            "G6": "Model Variant",
            "F7": "Speed",
            "F8": "Capacity",
            "F13": "Shaft Type",
            "F16": "Car Type",
            "F18": "Door Opening Type",
            "C25": "Shaft Door Material",
            "F20": "Car Door Material",
            "F25": "Shaft Door Material",
            "C21": "Ceiling",
            "D22": "Mirror Height",
            "G22": "Handrail",
            "C23": "COP",
            "F23": "COP Button",
            "C26": "LOP",
            "F26": "LOP",
            "C27": "LIP",
            "F27": "LIP",
            "C22": "Mirror Wall",
            "F22": "Handrail Position",
            "C10": "Control System",
            "C30": "Option Yes No",
            "C18": "Fire Rating",
            "C16": "Car Height",
            "C28": "Forced Yes",
            "C48": "Air Conditioner",
            "F17": "Door Height",
            "C20": "Car Wall Material",
            "C49": "RCC",
            "F21": "Floor",
            "C19": "Cabin Design",
        },
    )


def parse_lehy_base(wb) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for sheet in ["Basic price(LEHY-III)", "Basic price(LEHY-L-PRO)", "Basic price(ELENESSA)"]:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        current_series = None
        current_speed = None
        current_stops = None
        for row in rows[3:]:
            series = clean(row[0] if len(row) > 0 else None) or current_series
            capacity_text = clean(row[1] if len(row) > 1 else None)
            speed_text = clean(row[2] if len(row) > 2 else None) or current_speed
            stops = clean(row[3] if len(row) > 3 else None) or current_stops
            if clean(row[0] if len(row) > 0 else None):
                current_series = series
            if clean(row[2] if len(row) > 2 else None):
                current_speed = speed_text
            if clean(row[3] if len(row) > 3 else None):
                current_stops = stops
            if not series or not capacity_text:
                continue
            capacity = int(as_number(str(capacity_text).replace("kg", "")) or 0)
            speed = as_number(str(speed_text).replace("m/s", "")) or 0
            entries.append(
                {
                    "series": str(series).replace("\n", " "),
                    "capacity": capacity,
                    "speed": speed,
                    "basicStops": int(as_number(stops) or 0),
                    "basicPrice": clean(row[4] if len(row) > 4 else None),
                    "pricePerStop": clean(row[5] if len(row) > 5 else None),
                    "overHeightPer1000": clean(row[6] if len(row) > 6 else None),
                    "pricePerDoor2D2G": clean(row[7] if len(row) > 7 else None),
                }
            )
    return entries


def parse_lehy_functions(wb) -> list[dict[str, Any]]:
    ws = wb["Functions"]
    entries: list[dict[str, Any]] = []
    for row in list(ws.iter_rows(values_only=True))[2:]:
        for col in [0, 3]:
            code = clean(row[col] if col < len(row) else None)
            price = clean(row[col + 1] if col + 1 < len(row) else None)
            if not code:
                continue
            entries.append({"code": str(code), "price": price})
    return entries


SMEC_MATERIAL_LABELS = {
    "SUS-H": "Hairline SUS",
    "SUS-M": "Mirror SUS",
    "SUS-I": "Irregularly-lined SUS",
    "SUS-S": "Sand pattern SUS",
}

SMEC_CATALOG_DESCRIPTIONS = {
    "MELD(ELENESSA)": "Аварийное устройство эвакуации при отключении основного питания.",
    "ELD(LEHY)": "При отключении питания перемещает кабину на ближайший этаж, выравнивает ее и открывает двери.",
    "MFP": "При отсутствии вызовов кабина возвращается на основной этаж и остается на нем.",
    "AIL": "После регистрации вызова выбирается оптимальная кабина; пассажиру сообщается назначенный лифт.",
    "NS-CB": "Отмена обслуживания выбранных этажей кнопками панели управления и установочным переключателем.",
    "AUTL": "Этажный индикатор автоматического режима работы лифта.",
    "BA": "Передача основных сигналов рабочего состояния лифта во внешнюю систему.",
    "CLO-A": "Автоматическое отключение освещения кабины при длительном отсутствии вызовов.",
    "DPS": "В заданное время затишья кабины направляются на верхний этаж по требованиям движения.",
    "DKO-TB": "Удержание кнопки увеличивает время нахождения дверей в открытом состоянии.",
    "RCS": "Дистанционный запуск и остановка кабины внешним переключателем.",
    "SCS-B": "Блокировка выбранных этажей паролем на панели управления кабины.",
    "OHS": "Рассредоточенная стоянка кабин на главном и центральном этажах при групповом управлении.",
    "UPS": "В часы пик кабины непрерывно направляются на основной посадочный этаж.",
    "FHL": "Мигающий этажный фонарь показывает прибытие кабины и направление движения.",
    "MBS": "Многолучевая световая завеса повторно открывает двери при обнаружении пассажира или предмета.",
    "NDG": "Если двери долго остаются открытыми, включается звуковой сигнал и выполняется попытка закрывания.",
}

SMEC_TITANIUM_CODES = [
    "001",
    "002",
    "003",
    "004",
    "005",
    "006",
    "500",
    "501",
    "502",
    "503",
    "504",
    "505",
    "506",
]


def parse_capacity(value: Any) -> int | None:
    number = as_number(str(value).replace("kg", "")) if value is not None else None
    return int(number) if number is not None else None


def smec_capacity_columns(
    ws,
    header_row: int,
    variant_row: int | None = None,
    start_column: int = 4,
    end_column: int = 25,
) -> list[tuple[int, int, str | None]]:
    columns: list[tuple[int, int, str | None]] = []
    current_capacity: int | None = None
    for column in range(start_column, end_column + 1):
        capacity = parse_capacity(ws.cell(header_row, column).value)
        if capacity is not None:
            current_capacity = capacity
        elif variant_row is None:
            continue

        variant = clean(ws.cell(variant_row, column).value) if variant_row is not None else None
        if current_capacity is not None and (variant_row is None or variant is not None):
            columns.append((column, current_capacity, str(variant) if variant is not None else None))
    return columns


def smec_titanium_row(ws, start_row: int, end_row: int, titanium_code: str) -> int | None:
    if titanium_code == "002":
        marker = "ZDT-002"
    elif titanium_code == "500":
        marker = "ZDT-500"
    elif titanium_code == "502":
        marker = "ZDT-502"
    elif titanium_code in {"001", "003", "004", "005", "006"}:
        marker = "ZDT-001,003"
    else:
        marker = "ZDT-501,503"

    for row in range(start_row, end_row + 1):
        value = str(clean(ws.cell(row, 3).value) or "")
        if marker in value:
            return row
    return None


def append_smec_material_prices(
    entries: list[dict[str, Any]],
    ws,
    category: str,
    columns: list[tuple[int, int, str | None]],
    base_start_row: int,
    base_end_row: int,
    etching_row: int | None,
    titanium_start_row: int,
    titanium_end_row: int,
) -> None:
    base_rows: dict[str, int] = {}
    for row in range(base_start_row, base_end_row + 1):
        label = str(clean(ws.cell(row, 3).value) or "")
        if label:
            base_rows[label] = row

    for column, capacity, variant in columns:
        for material_code, material_label in SMEC_MATERIAL_LABELS.items():
            base_row = base_rows.get(material_label)
            base_price = as_number(ws.cell(base_row, column).value) if base_row else None
            if base_price is not None:
                entries.append(
                    {
                        "category": category,
                        "code": material_code,
                        "capacity": capacity,
                        "variant": variant,
                        "price": base_price,
                    }
                )

            for titanium_code in SMEC_TITANIUM_CODES:
                titanium_row = smec_titanium_row(
                    ws,
                    titanium_start_row,
                    titanium_end_row,
                    titanium_code,
                )
                etching_price = (
                    as_number(ws.cell(etching_row, column).value)
                    if etching_row is not None
                    else 0
                )
                titanium_price = (
                    as_number(ws.cell(titanium_row, column).value)
                    if titanium_row is not None
                    else None
                )
                if base_price is None or etching_price is None or titanium_price is None:
                    continue
                entries.append(
                    {
                        "category": category,
                        "code": f"ZDT-{titanium_code} {material_code}",
                        "capacity": capacity,
                        "variant": variant,
                        "price": base_price + etching_price + titanium_price,
                    }
                )


def append_smec_matrix_prices(
    entries: list[dict[str, Any]],
    ws,
    category: str,
    columns: list[tuple[int, int, str | None]],
    start_row: int,
    end_row: int,
    code_column: int = 2,
    code_aliases: dict[str, str] | None = None,
) -> None:
    aliases = code_aliases or {}
    for row in range(start_row, end_row + 1):
        raw_code = clean(ws.cell(row, code_column).value)
        if raw_code is None:
            continue
        code = aliases.get(str(raw_code).strip(), str(raw_code).strip())
        for column, capacity, variant in columns:
            price = as_number(ws.cell(row, column).value)
            if price is None:
                continue
            entries.append(
                {
                    "category": category,
                    "code": code,
                    "capacity": capacity,
                    "variant": variant,
                    "price": price,
                }
            )


def parse_smec_decorations(wb) -> list[dict[str, Any]]:
    ws = wb["Decoration"]
    entries: list[dict[str, Any]] = []

    cage_columns = smec_capacity_columns(ws, 5, 6)
    design_columns = smec_capacity_columns(ws, 22, 23)
    cabin_columns = smec_capacity_columns(ws, 39, end_column=18)
    finish_columns = smec_capacity_columns(ws, 65, end_column=18)
    landing_columns = smec_capacity_columns(ws, 105, end_column=18)

    append_smec_matrix_prices(entries, ws, "CarDesign", design_columns, 24, 37)
    append_smec_material_prices(entries, ws, "CarWall", cage_columns, 7, 14, 15, 16, 20)
    append_smec_material_prices(entries, ws, "FrontPanel", cabin_columns, 40, 43, 44, 45, 49)
    append_smec_material_prices(entries, ws, "CarDoor", cabin_columns, 50, 57, 58, 59, 63)
    append_smec_matrix_prices(entries, ws, "Ceiling", finish_columns, 66, 79)
    append_smec_matrix_prices(
        entries,
        ws,
        "Floor",
        finish_columns,
        80,
        86,
        code_aliases={
            "Car floor concave-down (floor by local)": "concave-down",
        },
    )
    append_smec_matrix_prices(entries, ws, "Mirror", finish_columns, 87, 88)
    append_smec_matrix_prices(entries, ws, "Handrail", finish_columns, 89, 98)
    append_smec_material_prices(entries, ws, "LandingDoor", landing_columns, 106, 113, 114, 115, 119)
    append_smec_matrix_prices(
        entries,
        ws,
        "DoorAddon",
        landing_columns,
        120,
        124,
        code_column=2,
        code_aliases={
            "Firerated door(E60)": "E120",
            "Firerated door(EI60)": "EI120",
            "Glass door(ZPKG-200)": "ZPKG-200",
            "Glass door(ZPKG-150)": "ZPKG-150",
            "Glass door(ZPKG-050)": "ZPKG-050",
        },
    )

    jamb_columns = [
        (column, 0, str(clean(ws.cell(126, column).value) or ""))
        for column in range(4, 8)
        if clean(ws.cell(126, column).value)
    ]
    append_smec_material_prices(entries, ws, "Jamb", jamb_columns, 127, 132, None, 133, 137)
    append_smec_faceplate_prices(entries, ws, "CopFaceplate", 141, 152)
    append_smec_faceplate_prices(entries, ws, "LopFaceplate", 153, 164)
    return entries


def append_smec_faceplate_prices(
    entries: list[dict[str, Any]],
    ws,
    category: str,
    start_row: int,
    end_row: int,
) -> None:
    titanium_groups = {
        "002": ["002"],
        "001,003-006": ["001", "003", "004", "005", "006", "007"],
        "001,003~006": ["001", "003", "004", "005", "006", "007"],
        "500": ["500"],
        "502": ["502"],
        "501,503-506": ["501", "503", "504", "505", "506"],
        "501,503~506": ["501", "503", "504", "505", "506"],
    }
    for row in range(start_row, end_row + 1):
        label = str(clean(ws.cell(row, 2).value) or "")
        price = as_number(ws.cell(row, 4).value)
        if not label or price is None:
            continue

        material = "SUS-M" if "Mirror SUS" in label else "SUS-H"
        match = re.search(r"ZDT-([0-9,~\\-]+)", label)
        if match is None:
            entries.append(
                {
                    "category": category,
                    "code": material,
                    "capacity": 0,
                    "variant": None,
                    "price": price,
                }
            )
            continue

        for titanium in titanium_groups.get(match.group(1), [match.group(1)]):
            entries.append(
                {
                    "category": category,
                    "code": f"ZDT-{titanium} {material}",
                    "capacity": 0,
                    "variant": None,
                    "price": price,
                }
            )


def parse_smec_option_descriptions(wb) -> dict[str, dict[str, Any]]:
    ws = wb.active
    descriptions: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(values_only=True):
        code = clean(row[1] if len(row) > 1 else None)
        description = clean(row[2] if len(row) > 2 else None)
        is_standard = clean(row[3] if len(row) > 3 else None)
        if not code:
            continue
        descriptions[str(code)] = {
            "description": description,
            "isStandard": bool(as_number(is_standard) or 0),
        }
    return descriptions


def parse_smec_car_designs(wb, image_map: dict[str, str]) -> list[dict[str, Any]]:
    ws = wb.active
    entries: list[dict[str, Any]] = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        code = clean(row[0] if len(row) > 0 else None)
        if not code:
            continue
        entries.append(
            {
                "code": str(code),
                "wallDescription": clean(row[1] if len(row) > 1 else None),
                "doorDescription": clean(row[2] if len(row) > 2 else None),
                "lastParagraph": clean(row[3] if len(row) > 3 else None),
                "imageUrl": image_map.get(str(code)),
            }
        )
    return entries


def merge_smec_car_design_choices(entries: list[dict[str, Any]], choices: list[str]) -> list[dict[str, Any]]:
    by_code = {entry["code"]: entry for entry in entries}
    merged: list[dict[str, Any]] = []
    for code in choices:
        if code in by_code:
            merged.append(by_code.pop(code))
            continue
        merged.append(
            {
                "code": code,
                "wallDescription": "Индивидуальный дизайн кабины" if code == "Customized" else "Car design",
                "doorDescription": None,
                "lastParagraph": None,
                "imageUrl": None,
            }
        )
    merged.extend(by_code.values())
    return merged


def parse_smec_power(wb) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            capacity = as_number(row[0] if len(row) > 0 else None)
            speed = as_number(row[1] if len(row) > 1 else None)
            power = as_number(row[2] if len(row) > 2 else None)
            if capacity is None or speed is None or power is None:
                continue
            entries.append(
                {
                    "series": sheet,
                    "capacity": int(capacity),
                    "speed": speed,
                    "power": power,
                }
            )
    return entries


def parse_smec_spec_fields(wb) -> list[dict[str, Any]]:
    ws = wb.active
    current_group = "GENERAL"
    entries: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        label = clean(row[0] if len(row) > 0 else None)
        group = clean(row[2] if len(row) > 2 else None)
        if group:
            current_group = str(group)
        if not label:
            continue
        entries.append(
            {
                "group": current_group,
                "label": str(label),
                "row": row_index,
            }
        )
    return entries


def parse_smec_choice_groups(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_name = "ЛC5" if "ЛC5" in workbook.sheetnames else "P1" if "P1" in workbook.sheetnames else workbook.sheetnames[0]
    return read_validation_choices(
        path,
        sheet_name,
        {
        "E3": "Project Type",
        "F3": "Project Type",
        "C5": "Ele Series",
        "E5": "Ele Type",
        "F5": "Ele Type",
        "G5": "Ele Type",
        "H5": "Ele Type",
        "E6": "Manufacturing Standard",
        "F6": "Manufacturing Standard",
        "G6": "Manufacturing Standard",
        "H6": "Manufacturing Standard",
        "H7": "Operation",
        "C8": "Capacity",
        "E8": "Speed",
        "H12": "Shaft Door Type",
        "F14": "Door Mode",
        "C16": "Car Design",
        "D16": "Car Design",
        "C17": "Ceiling",
        "D17": "Ceiling",
        "F17": "Floor Type",
        "H17": "Floor Pattern",
        "C19": "Mirror",
        "D19": "Mirror",
        "G19": "Mirror Position",
        "H19": "Mirror Position",
        "C20": "COP",
        "D20": "COP",
        "F20": "COP 2",
        "H20": "COP Button",
        "C21": "Wheelchair COP",
        "D21": "Wheelchair COP",
        "F21": "Wheelchair COP 2",
        "H21": "Wheelchair COP Button",
        "C23": "Jamb",
        "F23": "Jamb",
        "C24": "Sill Bracket",
        "D24": "Sill Bracket",
        "E24": "Sill Bracket",
        "F24": "Sill Bracket",
        "G24": "Sill Bracket",
        "H24": "Sill Bracket",
        "C26": "LOP",
        "F26": "LOP",
        "E26": "LOP Button",
        "H26": "LOP Button",
        "C27": "Auxiliary LOP",
        "F27": "Auxiliary LOP",
        "E27": "Auxiliary LOP Button",
        "H27": "Auxiliary LOP Button",
        "C28": "Hall Indicator",
        "D28": "Hall Indicator",
        "E28": "Hall Indicator",
        "F28": "Hall Indicator",
        "G28": "Hall Indicator",
        "H28": "Hall Indicator",
        "C29": "Hall Lantern",
        "D29": "Hall Lantern",
        "E29": "Hall Lantern",
        "F29": "Hall Lantern",
        "G29": "Hall Lantern",
        "H29": "Hall Lantern",
        },
    )


def parse_smec_model_groups(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheet_name = "ЛC5" if "ЛC5" in workbook.sheetnames else "P1" if "P1" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    allowed_series = {"LEHY Series", "ELENESSA", "Panoramic"}
    groups: list[dict[str, Any]] = []

    for row in range(9, 15):
        series = clean(worksheet.cell(row, 17).value)
        if not series:
            continue
        series_name = str(series).strip()
        if series_name not in allowed_series:
            continue

        options: list[str] = []
        cells: list[str] = []
        for column in range(18, worksheet.max_column + 1):
            value = clean(worksheet.cell(row, column).value)
            if value is None or is_excel_stop_marker(value):
                break
            options.append(str(value))
            cells.append(worksheet.cell(row, column).coordinate)

        groups.append(
            {
                "name": f"Ele Type: {series_name}",
                "sourceSheet": sheet_name,
                "cells": cells,
                "options": options,
            }
        )

    return groups


def copy_smec_images() -> tuple[list[dict[str, Any]], dict[str, str]]:
    source_dir = ROOT / "SMEC"
    entries: list[dict[str, Any]] = []
    image_map: dict[str, str] = {}
    if not source_dir.exists():
        return entries, image_map

    SMEC_ASSET_OUT.mkdir(parents=True, exist_ok=True)
    for source in sorted(source_dir.glob("Pic_*.png")):
        code = source.stem.removeprefix("Pic_")
        target_name = source.name.replace(" ", "_")
        target = SMEC_ASSET_OUT / target_name
        shutil.copy2(source, target)
        url = f"/assets/smec/{target_name}"
        image_map[code] = url
        image_map[code.replace("-", "■-", 1)] = url
        image_map[code.replace("■", "")] = url
        entries.append({"code": code, "imageUrl": url})
    return entries, image_map


def enrich_entries(
    entries: list[dict[str, Any]],
    descriptions: dict[str, dict[str, Any]],
    image_map: dict[str, str],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for entry in entries:
        code = str(entry.get("code") or "")
        description = descriptions.get(code, {})
        enriched.append(
            {
                **entry,
                "description": description.get("description") or SMEC_CATALOG_DESCRIPTIONS.get(code),
                "isStandard": description.get("isStandard"),
                "imageUrl": image_map.get(code),
            }
        )
    return enriched


def enrich_visual_entries(
    entries: list[dict[str, Any]],
    descriptions: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for entry in entries:
        code = str(entry.get("code") or "")
        description = descriptions.get(code, {})
        enriched.append(
            {
                **entry,
                "description": description.get("description"),
            }
        )
    return enriched


def parse_smec_floor_patterns(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet_name = "ЛC5" if "ЛC5" in workbook.sheetnames else workbook.sheetnames[0]
    ws = workbook[sheet_name]
    floor_types = []
    for group in read_validation_choices(path, sheet_name, {"F17": "Floor Type"}):
        if group["name"] == "Floor Type":
            floor_types = group["options"]
            break

    start_row = 21
    groups: list[dict[str, Any]] = []
    for index, floor_type in enumerate(floor_types):
        row = start_row + index
        options: list[str] = []
        for column in range(18, 28):
            value = clean(ws.cell(row, column).value)
            if is_excel_stop_marker(value):
                break
            if value is not None:
                options.append(str(value))
        groups.append({"floorType": floor_type, "options": options})
    return groups


def parse_lehy_group_control(wb) -> list[dict[str, Any]]:
    ws = wb["Control & Display"]
    entries: list[dict[str, Any]] = []
    for row in list(ws.iter_rows(values_only=True))[2:]:
        for col in [0, 4]:
            code = clean(row[col] if col < len(row) else None)
            price = clean(row[col + 1] if col + 1 < len(row) else None)
            if code and as_number(price) is not None:
                entries.append({"code": str(code), "price": price})
    return entries


def expand_control_codes(value: Any) -> list[str]:
    code = str(clean(value) or "").strip()
    if not code:
        return []
    if "/" not in code:
        return [code]

    first, *suffixes = code.split("/")
    prefix_match = re.match(r"^(.*?[-■●])([^/]+)$", first)
    if prefix_match is None:
        return [part.strip() for part in code.split("/") if part.strip()]

    prefix, first_suffix = prefix_match.groups()
    return [prefix + first_suffix] + [
        suffix.strip() if "-" in suffix else prefix + suffix.strip()
        for suffix in suffixes
        if suffix.strip()
    ]


def parse_smec_control_prices(wb) -> list[dict[str, Any]]:
    ws = wb["Control & Display"]
    entries: list[dict[str, Any]] = []

    def append(category: str, row: int, code_column: int, price_column: int) -> None:
        price = clean(ws.cell(row, price_column).value)
        if as_number(price) is None:
            return
        for code in expand_control_codes(ws.cell(row, code_column).value):
            entries.append({"category": category, "code": code, "price": price})

    for row in range(16, 31):
        append("HallIndicator", row, 2, 3)
        append("HallButton" if row <= 21 else "HallLantern", row, 6, 7)
    for row in range(33, 45):
        append("COP", row, 2, 3)
        append("COP2", row, 6, 7)
    for row in range(45, 49):
        append("WheelchairCOP", row, 2, 3)

    button_rows = {
        52: ["A51", "A54", "A61", "A64"],
        53: ["A71"],
        54: ["A23"],
        55: ["A27"],
    }
    for row, codes in button_rows.items():
        price = clean(ws.cell(row, 3).value)
        for code in codes:
            entries.append({"category": "Button", "code": code, "price": price})
    return entries


def parse_smec_cwt_prices(wb) -> list[dict[str, Any]]:
    ws = wb["Functions"]
    entries: list[dict[str, Any]] = []
    for row in range(36, 42):
        for label_column, price_column, series in [
            (1, 2, "ELENESSA"),
            (4, 5, "LEHY"),
        ]:
            label = clean(ws.cell(row, label_column).value)
            price = as_number(ws.cell(row, price_column).value)
            if label is None or price is None:
                continue
            capacities = [int(value) for value in re.findall(r"\d+", str(label))]
            if not capacities:
                continue
            entries.append(
                {
                    "series": series,
                    "minCapacity": min(capacities),
                    "maxCapacity": max(capacities),
                    "price": price,
                }
            )
    return entries


def parse_containers(wb) -> list[dict[str, Any]]:
    ws = wb["List"]
    entries: list[dict[str, Any]] = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        capacity = as_number(row[0] if len(row) > 0 else None)
        stops = as_number(row[1] if len(row) > 1 else None)
        if capacity is None or stops is None:
            continue
        entries.append(
            {
                "capacity": int(capacity),
                "stops": int(stops),
                "efs": clean(row[2] if len(row) > 2 else None),
                "e312": clean(row[3] if len(row) > 3 else None),
                "single": clean(row[4] if len(row) > 4 else None),
                "mix": clean(row[5] if len(row) > 5 else None),
                "mixUnit": clean(row[6] if len(row) > 6 else None),
            }
        )
    return entries


def main() -> None:
    xizi = load_workbook(ROOT / "XIZI_price_list_20_02_2026.xlsx", data_only=True, read_only=False)
    lehy = load_workbook(
        ROOT / "Price_of_FOB_CNY__165__2025__20420__21521__Including_LEHY-L-Pro.xlsx",
        data_only=True,
        read_only=False,
    )
    containers = load_workbook(ROOT / "Containers.xlsx", data_only=True, read_only=False)
    spec_template = load_workbook(ROOT / "Spetsifikatsia_TKP" / "SPEC_TKP_ELE_v2.0.xlsx", data_only=True, read_only=False)
    cardesign = load_workbook(ROOT / "Spetsifikatsia_TKP" / "CARDESIGN.xlsx", data_only=True, read_only=False)
    options = load_workbook(ROOT / "Spetsifikatsia_TKP" / "OPTIONS.xlsx", data_only=True, read_only=False)
    power = load_workbook(ROOT / "Spetsifikatsia_TKP" / "POWER.xlsx", data_only=True, read_only=False)

    smec_visual_items, smec_image_map = copy_smec_images()
    xizi_visual_items, xizi_image_map, xizi_description_map = copy_xizi_finish_images()
    smec_option_descriptions = parse_smec_option_descriptions(options)
    xizi_option_descriptions = parse_xizi_option_descriptions(XIZI_FORM_EXAMPLE)
    xizi_decorations = enrich_xizi_visual_entries(
        parse_xizi_decoration(xizi),
        xizi_image_map,
        xizi_description_map,
    )

    smec_choice_groups = parse_smec_choice_groups(SMEC_FORM_EXAMPLE)
    smec_model_groups = parse_smec_model_groups(SMEC_FORM_EXAMPLE)
    smec_choice_groups.extend(smec_model_groups)
    smec_series = [
        model
        for group in smec_model_groups
        for model in group["options"]
    ]
    smec_car_design_choices = next(
        (group["options"] for group in smec_choice_groups if group["name"] == "Car Design"),
        [],
    )
    smec_car_designs = merge_smec_car_design_choices(
        parse_smec_car_designs(cardesign, smec_image_map),
        smec_car_design_choices,
    )

    catalog = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "currency": "CNY",
        "xizi": {
            "series": ["UN-Victor R", "UN-Victor MRL", "G3"],
            "basePrices": parse_xizi_base(xizi),
            "doors": parse_xizi_doors(xizi),
            "decorations": xizi_decorations,
            "options": enrich_xizi_options(
                parse_xizi_simple_price_sheet(xizi, "Options", "Options"),
                xizi_option_descriptions,
            ),
            "localRequirements": parse_xizi_simple_price_sheet(xizi, "LMR", "LMR"),
            "containers": parse_xizi_containers(xizi),
            "finishDescriptions": [],
            "visualItems": xizi_visual_items,
            "choiceGroups": parse_xizi_choice_groups(XIZI_FORM_EXAMPLE),
        },
        "smec": {
            "series": smec_series,
            "basePrices": parse_lehy_base(lehy),
            "decorations": parse_smec_decorations(lehy),
            "functions": enrich_entries(parse_lehy_functions(lehy), smec_option_descriptions, smec_image_map),
            "groupControl": enrich_entries(parse_lehy_group_control(lehy), smec_option_descriptions, smec_image_map),
            "controlPrices": parse_smec_control_prices(lehy),
            "cwtPrices": parse_smec_cwt_prices(lehy),
            "containers": parse_containers(containers),
            "carDesigns": smec_car_designs,
            "visualItems": enrich_visual_entries(smec_visual_items, smec_option_descriptions),
            "power": parse_smec_power(power),
            "specFields": parse_smec_spec_fields(spec_template),
            "choiceGroups": smec_choice_groups,
            "floorPatterns": parse_smec_floor_patterns(SMEC_FORM_EXAMPLE),
        },
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
