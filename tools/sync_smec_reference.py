#!/usr/bin/env python3
"""Synchronize SMEC price tables from the supplier workbook into the app catalog."""

from __future__ import annotations

import argparse
import json
import re
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SOURCE = Path("/Users/lehjke/Downloads/Price FOB CNY.xlsx")
DEFAULT_CATALOG = Path("src/TFlexDrawingService.Api/Data/pricing-catalog.json")
MATERIALS = ("SUS-H", "SUS-M", "SUS-I", "SUS-S")
TITANIUM_CODES = ("001", "002", "003", "004", "005", "006", "500", "501", "502", "503", "504", "505", "506")


def clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def number(value: Any) -> float | None:
    value = clean(value)
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    match = re.search(r"-?\d+(?:[.,]\d+)?", str(value))
    return float(match.group().replace(",", ".")) if match else None


def json_number(value: Any) -> int | float | None:
    parsed = number(value)
    if parsed is None:
        return None
    parsed = round(parsed, 6)
    return int(parsed) if parsed.is_integer() else parsed


def normalize_series(value: str) -> str:
    value = value.replace("\n", " ").strip()
    if value == "ELENESSA":
        return "ELENESSA (GQXL3M3)"
    return re.sub(r"(?<!\s)\(", " (", value)


def parse_base_prices(workbook) -> list[dict[str, Any]]:
    worksheet = workbook["Basic price"]
    entries: list[dict[str, Any]] = []
    series = ""
    speed: int | float | None = None
    basic_stops: int | None = None
    for row in range(4, worksheet.max_row + 1):
        if clean(worksheet.cell(row, 1).value):
            series = normalize_series(str(worksheet.cell(row, 1).value))
        if clean(worksheet.cell(row, 3).value):
            speed = json_number(worksheet.cell(row, 3).value)
        if clean(worksheet.cell(row, 4).value) is not None:
            basic_stops = int(number(worksheet.cell(row, 4).value) or 0)
        capacity = json_number(worksheet.cell(row, 2).value)
        basic_price = json_number(worksheet.cell(row, 5).value)
        if not series or capacity is None or speed is None or basic_stops is None or basic_price is None or basic_price < 0:
            continue
        if series == "LEHY-MRL-II":
            continue
        entries.append({
            "series": series,
            "capacity": int(capacity),
            "speed": speed,
            "basicStops": basic_stops,
            "basicPrice": basic_price,
            "pricePerStop": json_number(worksheet.cell(row, 6).value),
            "overHeightPer1000": json_number(worksheet.cell(row, 7).value),
            "pricePerDoor2D2G": json_number(worksheet.cell(row, 8).value),
        })
    return entries


def capacity_columns(worksheet, header_row: int, variant_row: int | None = None, end_column: int = 25):
    result = []
    current_capacity: int | None = None
    for column in range(4, end_column + 1):
        parsed = json_number(worksheet.cell(header_row, column).value)
        if parsed is not None:
            current_capacity = int(parsed)
        elif variant_row is None:
            continue
        variant = clean(worksheet.cell(variant_row, column).value) if variant_row else None
        if current_capacity is not None and (variant_row is None or variant is not None):
            result.append((column, current_capacity, str(variant) if variant is not None else None))
    return result


def append_matrix(entries, worksheet, category, columns, start_row, end_row, code_column=2, aliases=None):
    aliases = aliases or {}
    for row in range(start_row, end_row + 1):
        raw_code = clean(worksheet.cell(row, code_column).value)
        if raw_code is None:
            continue
        code = aliases.get(str(raw_code), str(raw_code))
        for column, capacity, variant in columns:
            price = json_number(worksheet.cell(row, column).value)
            if price is not None:
                entries.append({"category": category, "code": code, "capacity": capacity, "variant": variant, "price": price})


def append_materials(entries, worksheet, category, columns, base_rows, etching_row, titanium_rows):
    for column, capacity, variant in columns:
        etching = json_number(worksheet.cell(etching_row, column).value) or 0
        for material, base_row in zip(MATERIALS, base_rows):
            base = json_number(worksheet.cell(base_row, column).value)
            if base is None:
                continue
            entries.append({"category": category, "code": material, "capacity": capacity, "variant": variant, "price": base})
            for titanium, titanium_row in zip(TITANIUM_CODES, titanium_rows):
                extra = json_number(worksheet.cell(titanium_row, column).value)
                if extra is not None:
                    entries.append({
                        "category": category,
                        "code": f"ZDT-{titanium} {material}",
                        "capacity": capacity,
                        "variant": variant,
                        "price": base + etching + extra,
                    })


def parse_decorations(workbook) -> list[dict[str, Any]]:
    worksheet = workbook["Decoration"]
    entries: list[dict[str, Any]] = []
    cage = capacity_columns(worksheet, 4, 5)
    designs = capacity_columns(worksheet, 29, 30)
    cabin = capacity_columns(worksheet, 46, end_column=18)
    finishes = capacity_columns(worksheet, 92, end_column=18)
    landing = capacity_columns(worksheet, 136, end_column=18)
    jamb = [(column, 0, str(clean(worksheet.cell(165, column).value))) for column in range(4, 8)]

    append_materials(entries, worksheet, "CarWall", cage, range(10, 14), 14, range(15, 28))
    append_matrix(entries, worksheet, "CarDesign", designs, 31, 44)
    append_materials(entries, worksheet, "FrontPanel", cabin, range(51, 55), 55, range(56, 69))
    append_materials(entries, worksheet, "CarDoor", cabin, range(73, 77), 77, range(78, 91))
    append_matrix(entries, worksheet, "Ceiling", finishes, 93, 108)
    append_matrix(entries, worksheet, "Floor", finishes, 109, 115, aliases={"Car floor concave-down (floor by local)": "concave-down"})
    append_matrix(entries, worksheet, "Mirror", finishes, 116, 117)
    append_matrix(entries, worksheet, "Handrail", finishes, 118, 127)
    append_materials(entries, worksheet, "LandingDoor", landing, range(141, 145), 145, range(146, 159))
    append_matrix(entries, worksheet, "DoorAddon", landing, 159, 163)
    append_materials(entries, worksheet, "Jamb", jamb, range(170, 174), 174, range(175, 188))

    for category, start_row, end_row in (("CopFaceplate", 191, 218), ("LopFaceplate", 219, 246)):
        for row in range(start_row, end_row + 1):
            code = clean(worksheet.cell(row, 2).value)
            price = json_number(worksheet.cell(row, 4).value)
            if code is not None and price is not None:
                entries.append({"category": category, "code": str(code), "capacity": 0, "variant": None, "price": price})
    return entries


FUNCTION_CODES = {
    "AECH / stop": "AECH",
    "ITV / TR(m)": "ITV",
    "Decoration Weight / 100 kg": "Decoration Weight",
    "Steel nosing by seller\n(Sill Support) / stop": "Steel nosing by seller\n(Sill Support)",
    "HL ＞2400mm / 100 mm": "HL ＞2400mm",
    "HH ＞2100mm\n/ 100 mm / opening": "HH ＞2100mm",
    "AHC / stop": "AHC",
    "Roller guide shoes": "Roller guide shoe",
}


def normalized_code(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def parse_functions(workbook, existing: list[dict[str, Any]]) -> list[dict[str, Any]]:
    worksheet = workbook["Functions"]
    by_code = {normalized_code(item.get("code")): item for item in existing}
    entries = []
    for row in range(3, 68):
        raw_code = clean(worksheet.cell(row, 1).value)
        price = json_number(worksheet.cell(row, 2).value)
        if raw_code is None or price is None:
            continue
        code = FUNCTION_CODES.get(str(raw_code), str(raw_code))
        previous = by_code.get(normalized_code(code), {})
        entries.append({
            "code": code,
            "price": price,
            "description": previous.get("description"),
            "isStandard": previous.get("isStandard"),
            "imageUrl": previous.get("imageUrl"),
        })
    previous_cwt = by_code.get(normalized_code("CWT Safety Gear"), {})
    entries.append({
        "code": "CWT Safety Gear",
        "price": None,
        "description": previous_cwt.get("description") or "Ловители на противовесе.",
        "isStandard": previous_cwt.get("isStandard"),
        "imageUrl": previous_cwt.get("imageUrl"),
    })
    return entries


def parse_group_control(workbook, existing: list[dict[str, Any]]) -> list[dict[str, Any]]:
    worksheet = workbook["Control & Display"]
    by_code = {normalized_code(item.get("code")): item for item in existing}
    entries = []
    for row in range(4, 25):
        code = clean(worksheet.cell(row, 1).value)
        price = json_number(worksheet.cell(row, 2).value)
        if code is None or price is None:
            continue
        previous = by_code.get(normalized_code(code), {})
        entries.append({
            "code": str(code),
            "price": price,
            "description": previous.get("description"),
            "isStandard": previous.get("isStandard"),
            "imageUrl": previous.get("imageUrl"),
        })
    return entries


def parse_control_prices(workbook) -> list[dict[str, Any]]:
    worksheet = workbook["Control & Display"]
    sections = (
        ("HallIndicator", 27, 50),
        ("HallButton", 51, 59),
        ("HallLantern", 60, 65),
        ("COP", 68, 80),
        ("WheelchairCOP", 81, 84),
        ("COP2", 85, 97),
    )
    entries = []
    for category, start_row, end_row in sections:
        for row in range(start_row, end_row + 1):
            code = clean(worksheet.cell(row, 2).value)
            price = json_number(worksheet.cell(row, 3).value)
            if code is not None and price is not None:
                entries.append({"category": category, "code": str(code), "price": price})
    for row in range(99, 102):
        code = clean(worksheet.cell(row, 1).value)
        price = json_number(worksheet.cell(row, 3).value)
        if code is not None and price is not None:
            entries.append({"category": "Button", "code": str(code), "price": price})
    return entries


def parse_cwt_prices(workbook) -> list[dict[str, Any]]:
    worksheet = workbook["Functions"]
    entries = []
    for header_row, price_row, series in ((69, 70, "ELENESSA"), (72, 73, "LEHY")):
        values = []
        for column in range(4, worksheet.max_column + 1):
            capacity = json_number(worksheet.cell(header_row, column).value)
            price = json_number(worksheet.cell(price_row, column).value)
            if capacity is not None and price is not None:
                values.append((int(capacity), price))
        start = 0
        while start < len(values):
            end = start
            while end + 1 < len(values) and values[end + 1][1] == values[start][1]:
                end += 1
            entries.append({
                "series": series,
                "minCapacity": values[start][0],
                "maxCapacity": values[end][0],
                "price": values[start][1],
            })
            start = end + 1
    return entries


def synchronized_catalog(source: Path, catalog_path: Path, *, update_timestamp: bool) -> dict[str, Any]:
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    result = deepcopy(catalog)
    workbook = load_workbook(source, data_only=True, read_only=False)
    smec = result["smec"]
    smec["basePrices"] = parse_base_prices(workbook)
    smec["decorations"] = parse_decorations(workbook)
    smec["functions"] = parse_functions(workbook, smec.get("functions", []))
    smec["groupControl"] = parse_group_control(workbook, smec.get("groupControl", []))
    smec["controlPrices"] = parse_control_prices(workbook)
    smec["cwtPrices"] = parse_cwt_prices(workbook)
    if update_timestamp:
        result["generatedAt"] = datetime.now(timezone.utc).isoformat()
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--catalog", type=Path, default=DEFAULT_CATALOG)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    result = synchronized_catalog(args.source, args.catalog, update_timestamp=not args.check)
    serialized = json.dumps(result, ensure_ascii=False, indent=2) + "\n"
    if args.check:
        current = args.catalog.read_text(encoding="utf-8")
        if serialized != current:
            raise SystemExit("SMEC catalog differs from the supplier workbook")
        print("SMEC catalog matches the supplier workbook")
        return
    args.catalog.write_text(serialized, encoding="utf-8")
    print(
        f"Updated {args.catalog}: "
        f"{len(result['smec']['basePrices'])} base rows, "
        f"{len(result['smec']['decorations'])} decoration rows"
    )


if __name__ == "__main__":
    main()
